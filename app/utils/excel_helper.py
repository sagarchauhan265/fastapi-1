from io import BytesIO
from fastapi import HTTPException
import openpyxl

REQUIRED_COLUMNS = {"name", "description", "price", "stock", "unit", "is_active", "cat_id"}
MAX_ROWS = 500
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB


def parse_excel(file_bytes: bytes) -> list[dict]:
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="FILE_TOO_LARGE_MAX_5MB")

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="INVALID_EXCEL_FILE")

    ws = wb.active

    # Validate headers exist
    if ws.max_row < 1:
        raise HTTPException(status_code=400, detail="EXCEL_FILE_IS_EMPTY")

    headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in ws[1]]

    # Check for missing required columns
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"MISSING_COLUMNS: {', '.join(sorted(missing))}"
        )

    # Check for duplicate column headers
    non_empty_headers = [h for h in headers if h]
    if len(non_empty_headers) != len(set(non_empty_headers)):
        raise HTTPException(status_code=400, detail="DUPLICATE_COLUMN_HEADERS")

    all_rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Filter out fully empty rows silently
    data_rows = [
        (row_index, row)
        for row_index, row in enumerate(all_rows, start=2)
        if any(cell is not None and str(cell).strip() != "" for cell in row)
    ]

    if not data_rows:
        raise HTTPException(status_code=400, detail="EXCEL_FILE_IS_EMPTY")

    if len(data_rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"EXCEEDS_MAX_ROWS_LIMIT_{MAX_ROWS}")

    result = []
    for row_index, row in data_rows:
        row_dict = {"_row": row_index}
        for col_index, header in enumerate(headers):
            if header:
                row_dict[header] = row[col_index]
        result.append(row_dict)

    return result
